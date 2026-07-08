"""
report_generator module — P14 Excel / SQLite reporting (W8, W16).

Commands:
  export_excel         — Generate a multi-sheet .xlsx: Elements, Params, QA Findings.
  export_sqlite_summary — Write elements + types into a fresh SQLite db.
"""
from __future__ import annotations

import json
import logging
import sqlite3
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2C5F8A")

SEVERITY_COLORS = {
    "error": "FFE0E0",
    "warning": "FFF3CD",
    "info": "E0F0FF",
}


def _ws_dir(workspace: Any) -> Path:
    return workspace.allowed_directories[0]

def _get_data(snapshot_id: str, workspace: Any):
    if not snapshot_id:
        from revit_mcp_server.semantic.engine import generate_mock_snapshot
        snap = generate_mock_snapshot()
        return (
            [el.model_dump(by_alias=True) for el in snap.elements],
            [t.model_dump() for t in snap.types],
        )
    path = _ws_dir(workspace) / "snapshots" / f"{snapshot_id}.json"
    if not path.exists():
        raise ValueError(f"Snapshot '{snapshot_id}' not found.")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("elements", []), data.get("types", [])

def _match_element(el: Dict[str, Any], filter_dsl: Dict[str, Any]) -> bool:
    if not filter_dsl:
        return True
    for key, val in filter_dsl.items():
        if key == "category":
            cats = val if isinstance(val, list) else [val]
            if el.get("category") not in cats:
                return False
        elif key == "family" and el.get("family") != val:
            return False
        elif key == "type_name" and el.get("type_name") != val:
            return False
    return True

def _write_header(ws: Any, headers: List[str]) -> None:
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(title) + 2)

def _load_qaqc_issues(workspace: Any) -> List[Dict[str, Any]]:
    db_path = _ws_dir(workspace) / "qaqc_issues.db"
    if not db_path.exists():
        return []
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM issues").fetchall()
    conn.close()
    return [dict(r) for r in rows]


class ReportGeneratorModule:

    def export_excel(
        self,
        snapshot_id: str = "",
        element_filter: Dict[str, Any] = None,
        param_names: List[str] = None,
        include_qaqc: bool = True,
        output_filename: str = "model_report.xlsx",
        workspace: Any = None,
        **_,
    ) -> Dict[str, Any]:
        elements, types = _get_data(snapshot_id, workspace)
        filter_dsl = element_filter or {}
        param_names = param_names or []

        matched = [el for el in elements if _match_element(el, filter_dsl)]

        wb = openpyxl.Workbook()

        # --- Sheet 1: Elements ---
        ws_el = wb.active
        ws_el.title = "Elements"
        el_headers = ["UID", "Element ID", "Category", "Family", "Type Name", "Level UID", "Workset"]
        _write_header(ws_el, el_headers)
        for r, el in enumerate(matched, 2):
            ws_el.append([
                el.get("uid", ""),
                el.get("element_id", ""),
                el.get("category", ""),
                el.get("family", ""),
                el.get("type_name", ""),
                el.get("level_uid", ""),
                el.get("workset", ""),
            ])

        # --- Sheet 2: Parameters ---
        if param_names:
            ws_param = wb.create_sheet("Parameters")
            param_headers = ["UID", "Category", "Type Name"] + param_names
            _write_header(ws_param, param_headers)
            for r, el in enumerate(matched, 2):
                row = [el.get("uid", ""), el.get("category", ""), el.get("type_name", "")]
                for pname in param_names:
                    pinfo = (el.get("params", {}) or {}).get(pname)
                    row.append(pinfo.get("v") if pinfo else "")
                ws_param.append(row)

        # --- Sheet 3: Types ---
        ws_types = wb.create_sheet("Types")
        _write_header(ws_types, ["Family", "Type Name", "Category", "Family Source"])
        for t in types:
            ws_types.append([
                t.get("family", ""),
                t.get("type_name", ""),
                t.get("category", ""),
                t.get("family_source", "loadable"),
            ])

        # --- Sheet 4: QA/QC Issues (optional) ---
        if include_qaqc:
            ws_qa = wb.create_sheet("QA_QC Issues")
            qa_headers = ["Issue ID", "Rule ID", "Severity", "Element UID", "Label", "Message", "Status", "Created At"]
            _write_header(ws_qa, qa_headers)
            issues = _load_qaqc_issues(workspace)
            for r, issue in enumerate(issues, 2):
                ws_qa.append([
                    issue.get("id", ""),
                    issue.get("rule_id", ""),
                    issue.get("severity", ""),
                    issue.get("element_uid", ""),
                    issue.get("label", ""),
                    issue.get("message", ""),
                    issue.get("status", ""),
                    issue.get("created_at", ""),
                ])
                # Color row by severity
                sev = issue.get("severity", "")
                fill_color = SEVERITY_COLORS.get(sev)
                if fill_color:
                    fill = PatternFill("solid", fgColor=fill_color)
                    for col in range(1, len(qa_headers) + 1):
                        ws_qa.cell(row=r, column=col).fill = fill

        # --- Save ---
        out_path = _ws_dir(workspace) / output_filename
        wb.save(str(out_path))

        return {
            "status": "exported",
            "output_file": str(out_path),
            "element_count": len(matched),
            "type_count": len(types),
            "sheets": wb.sheetnames,
        }

    def export_sqlite_summary(
        self,
        snapshot_id: str = "",
        output_filename: str = "model_data.db",
        workspace: Any = None,
        **_,
    ) -> Dict[str, Any]:
        elements, types = _get_data(snapshot_id, workspace)
        out_path = _ws_dir(workspace) / output_filename

        conn = sqlite3.connect(str(out_path))
        conn.executescript("""
            DROP TABLE IF EXISTS elements;
            DROP TABLE IF EXISTS types;
            CREATE TABLE elements (
                uid TEXT PRIMARY KEY,
                element_id TEXT,
                category TEXT,
                family TEXT,
                type_name TEXT,
                level_uid TEXT,
                workset TEXT,
                host_uid TEXT
            );
            CREATE TABLE types (
                family TEXT,
                type_name TEXT,
                category TEXT,
                family_source TEXT
            );
        """)

        conn.executemany(
            "INSERT OR REPLACE INTO elements VALUES (?,?,?,?,?,?,?,?)",
            [
                (
                    el.get("uid", ""),
                    el.get("element_id", ""),
                    el.get("category", ""),
                    el.get("family", ""),
                    el.get("type_name", ""),
                    el.get("level_uid", ""),
                    el.get("workset", ""),
                    el.get("host_uid", ""),
                )
                for el in elements
            ]
        )

        conn.executemany(
            "INSERT INTO types VALUES (?,?,?,?)",
            [
                (
                    t.get("family", ""),
                    t.get("type_name", ""),
                    t.get("category", ""),
                    t.get("family_source", "loadable"),
                )
                for t in types
            ]
        )

        conn.commit()
        conn.close()

        return {
            "status": "exported",
            "output_file": str(out_path),
            "element_count": len(elements),
            "type_count": len(types),
        }
