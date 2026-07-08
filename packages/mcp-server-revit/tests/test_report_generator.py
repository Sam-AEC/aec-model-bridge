"""Tests for Phase 14 — Report Generator module (W8, W16)."""
import importlib.util
import json
import sqlite3
import pytest
from pathlib import Path
from revit_mcp_server.semantic.engine import generate_mock_snapshot


def _load_mod(relpath: str, name: str):
    p = Path(__file__).parent.parent / relpath
    spec = importlib.util.spec_from_file_location(name, p)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_rg = _load_mod("src/revit_mcp_server/modules/report_generator/module.py", "_rg_impl")
ReportGeneratorModule = _rg.ReportGeneratorModule


class MockWorkspace:
    def __init__(self, tmp_path: Path):
        self.allowed_directories = [tmp_path]


@pytest.fixture
def workspace(tmp_path):
    return MockWorkspace(tmp_path)


@pytest.fixture
def module():
    return ReportGeneratorModule()


def test_export_excel_creates_file(module, workspace, tmp_path):
    result = module.export_excel(
        param_names=["Mark", "FireRating"],
        include_qaqc=False,
        workspace=workspace,
    )
    assert result["status"] == "exported"
    out = Path(result["output_file"])
    assert out.exists()
    assert out.suffix == ".xlsx"
    assert result["element_count"] == 6


def test_export_excel_includes_params_sheet(module, workspace, tmp_path):
    import openpyxl
    result = module.export_excel(
        param_names=["Mark"],
        include_qaqc=False,
        workspace=workspace,
    )
    wb = openpyxl.load_workbook(result["output_file"])
    assert "Parameters" in wb.sheetnames
    ws = wb["Parameters"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "Mark" in headers


def test_export_excel_with_element_filter(module, workspace):
    result = module.export_excel(
        element_filter={"category": "OST_Doors"},
        include_qaqc=False,
        workspace=workspace,
    )
    assert result["element_count"] == 1


def test_export_excel_with_qaqc_no_issues(module, workspace):
    """QA sheet should exist even with no issue store."""
    import openpyxl
    result = module.export_excel(include_qaqc=True, workspace=workspace)
    wb = openpyxl.load_workbook(result["output_file"])
    assert "QA_QC Issues" in wb.sheetnames


def test_export_sqlite_creates_db(module, workspace, tmp_path):
    result = module.export_sqlite_summary(workspace=workspace)
    assert result["status"] == "exported"
    db_path = Path(result["output_file"])
    assert db_path.exists()
    
    conn = sqlite3.connect(str(db_path))
    count = conn.execute("SELECT COUNT(*) FROM elements").fetchone()[0]
    conn.close()
    assert count == 6


def test_export_sqlite_type_table(module, workspace):
    result = module.export_sqlite_summary(workspace=workspace)
    conn = sqlite3.connect(result["output_file"])
    types = conn.execute("SELECT * FROM types").fetchall()
    conn.close()
    assert len(types) >= 2


def test_export_excel_custom_filename(module, workspace, tmp_path):
    result = module.export_excel(
        output_filename="custom_report.xlsx",
        include_qaqc=False,
        workspace=workspace,
    )
    assert Path(result["output_file"]).name == "custom_report.xlsx"
